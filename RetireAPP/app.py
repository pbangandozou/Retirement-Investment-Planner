import io
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

# ==========================================
# PAGE CONFIGURATION
# ==========================================
st.set_page_config(
    page_title="Retirement & Investment Planner",
    page_icon="📈",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Inter:wght@400;500;600;700&display=swap');

.stApp { background:#f7f8f6; color:#17221d; }
.block-container { max-width:1180px; padding-top:2.4rem; padding-bottom:4rem; }
#MainMenu, footer, header { visibility:hidden; }

.brand-bar { display:flex; justify-content:space-between; align-items:center; padding-bottom:18px; margin-bottom:36px; border-bottom:1px solid #dfe5e1; }
.brand-name { font-size:15px; font-weight:700; letter-spacing:1px; color:#17221d; }
.brand-name span { color:#2f8053; }
.brand-meta,.section-meta,.sidebar-caption { font-family:'DM Mono',monospace; font-size:9px; letter-spacing:1px; color:#7c8982; text-transform:uppercase; }
.eyebrow { font-family:'DM Mono',monospace; font-size:9px; letter-spacing:1.5px; color:#2f8053; text-transform:uppercase; margin-bottom:8px; }
.page-title { font-size:38px; line-height:1.1; letter-spacing:-1.7px; font-weight:700; margin:0; color:#17221d; }
.page-subtitle { font-size:13px; line-height:1.7; color:#68756e; max-width:720px; margin-top:12px; }

.section-header { display:flex; justify-content:space-between; align-items:end; margin-top:38px; margin-bottom:16px; border-bottom:1px solid #dfe5e1; padding-bottom:12px; }
.section-title { font-size:17px; font-weight:600; letter-spacing:-.3px; color:#17221d; }

.kpi-card,.analysis-card { background:#fff; border:1px solid #d5ded8; padding:18px 20px; }
.kpi-card { min-height:112px; }
.kpi-label,.analysis-number { font-family:'DM Mono',monospace; font-size:9px; letter-spacing:1px; color:#7c8982; text-transform:uppercase; }
.kpi-value { font-family:'DM Mono',monospace; font-size:23px; font-weight:500; color:#17221d; margin-top:12px; }
.kpi-note { font-family:'DM Mono',monospace; font-size:9px; color:#2f8053; margin-top:5px; }
.analysis-title { font-size:15px; font-weight:600; margin-top:14px; color:#17221d; }
.analysis-text { font-size:11px; line-height:1.7; color:#68756e; margin-top:7px; }

section[data-testid='stSidebar'] { background:#edf2ee; border-right:1px solid #dfe5e1; }
.sidebar-brand { font-size:15px; font-weight:700; letter-spacing:1px; color:#17221d; margin-bottom:3px; }
.sidebar-brand span { color:#2f8053; }
.sidebar-section { font-family:'DM Mono',monospace; font-size:9px; letter-spacing:1.2px; color:#2f8053; text-transform:uppercase; margin-top:20px; margin-bottom:4px; }
.stNumberInput label,.stSlider label { font-size:11px !important; color:#536159 !important; }
.stNumberInput input { font-family:'DM Mono',monospace !important; font-size:11px !important; }
.stButton > button { width:100%; border-radius:4px; border:1px solid #17221d; background:#17221d; color:white; font-size:11px; font-weight:600; }
.stButton > button:hover { border-color:#2f8053; background:#2f8053; }
.stDownloadButton > button { width:100%; border-radius:4px; border:1px solid #2f8053; background:transparent; color:#2f8053; font-size:10px; font-weight:600; }
.stTabs [data-baseweb='tab-list'] { gap:0; border-bottom:1px solid #d5ded8; }
.stTabs [data-baseweb='tab'] { font-size:10px; color:#6e7b74; padding:12px 18px; }
.stTabs [aria-selected='true'] { color:#2f8053 !important; }
[data-testid='stDataFrame'] { border:1px solid #d5ded8; }
.app-footer { margin-top:55px; padding-top:18px; border-top:1px solid #dfe5e1; display:flex; justify-content:space-between; font-family:'DM Mono',monospace; font-size:8px; color:#849089; text-transform:uppercase; letter-spacing:.7px; }

@media (prefers-color-scheme: dark) {
 .stApp { background:#111613; color:#e8eee9; }
 .brand-bar,.section-header,.app-footer { border-color:#29332d; }
 .brand-name,.page-title,.section-title,.kpi-value,.analysis-title,.sidebar-brand { color:#e8eee9; }
 .brand-meta,.page-subtitle,.section-meta,.kpi-label,.analysis-text,.sidebar-caption { color:#9aa89f; }
 .kpi-card,.analysis-card { background:#171d19; border-color:#303b34; }
 section[data-testid='stSidebar'] { background:#171d19; border-color:#29332d; }
 .stTabs [data-baseweb='tab-list'],[data-testid='stDataFrame'] { border-color:#303b34; }
 .stTabs [data-baseweb='tab'] { color:#9aa89f; }
}
</style>

<div class="brand-bar">
  <div class="brand-name">Prustide Bangandozou<span>.</span></div>
  <div class="brand-meta">Finance · Analytics · Technology</div>
</div>

<div class="eyebrow">PERSONAL FINANCIAL MODEL · RETIREMENT PLANNING</div>
<div class="page-title">Retirement &amp; Personal Investment Planner</div>
<div class="page-subtitle">A long-term financial planning model for evaluating portfolio growth, inflation, savings behavior, retirement readiness, scenario analysis, and investment risk.</div>
""", unsafe_allow_html=True)


# ==========================================
# EXCEL GENERATOR FUNCTION
# ==========================================
def build_excel_in_memory(inputs):
  """Generates the full 7-tab Excel workbook dynamically based on Streamlit inputs."""
  wb = openpyxl.Workbook()
  tabs = [
      "01 — Dashboard",
      "02 — Assumptions",
      "03 — Accumulation",
      "04 — Retirement",
      "05 — Scenarios",
      "06 — Monte Carlo",
      "07 — Summary",
  ]

  wb.active.title = tabs[0]
  for tab in tabs[1:]:
    wb.create_sheet(title=tab)

  navy_header = PatternFill(
      start_color="002060", end_color="002060", fill_type="solid"
  )
  white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  bold_font = Font(name="Calibri", size=11, bold=True)
  thin_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )

  # Tab 2: Assumptions
  ws_assump = wb["02 — Assumptions"]
  ws_assump.views.sheetView[0].showGridLines = True
  ws_assump["A1"], ws_assump["B1"] = "Model Parameter", "Value"
  ws_assump["A1"].fill = ws_assump["B1"].fill = navy_header
  ws_assump["A1"].font = ws_assump["B1"].font = white_bold

  row_idx = 2
  for k, v in inputs.items():
    ws_assump.cell(row=row_idx, column=1, value=k.replace("_", " ").title())
    cell = ws_assump.cell(row=row_idx, column=2, value=v)
    if any(
        x in k
        for x in [
            "pct",
            "return",
            "inflation",
            "tax",
            "rate",
            "growth",
            "volatility",
        ]
    ):
      cell.number_format = "0.0%"
    elif any(
        x in k
        for x in ["portfolio", "income", "contribution", "spending", "security"]
    ):
      cell.number_format = "$#,##0"
    row_idx += 1

  # Tab 3: Accumulation
  ws_accum = wb["03 — Accumulation"]
  ws_accum.views.sheetView[0].showGridLines = True
  headers = [
      "Age",
      "Year",
      "Starting Balance",
      "User Contribution",
      "Employer Match",
      "Investment Gain",
      "Ending Portfolio",
      "Inflation-Adjusted",
  ]

  for col_num, h in enumerate(headers, 1):
    c = ws_accum.cell(row=1, column=col_num, value=h)
    c.fill, c.font, c.alignment = (
        navy_header,
        white_bold,
        Alignment(horizontal="center"),
    )

  years = inputs["retirement_age"] - inputs["current_age"]
  for i in range(1, years + 1):
    r = i + 1
    ws_accum.cell(row=r, column=1, value=f"='02 — Assumptions'!B2 + {i-1}")
    ws_accum.cell(row=r, column=2, value=i)
    ws_accum.cell(
        row=r,
        column=3,
        value="='02 — Assumptions'!B5" if i == 1 else f"=G{r-1}",
    )
    ws_accum.cell(
        row=r,
        column=4,
        value=(
            f"='02 — Assumptions'!B7 * ((1 + '02 — Assumptions'!B15) ^ (B{r}-1))"
        ),
    )
    ws_accum.cell(
        row=r,
        column=5,
        value=(
            "='02 — Assumptions'!B6 * '02 — Assumptions'!B8 * ((1 + '02 —"
            f" Assumptions'!B15) ^ (B{r}-1))"
        ),
    )
    ws_accum.cell(
        row=r,
        column=6,
        value=f"=(C{r} + (D{r}+E{r})/2) * '02 — Assumptions'!B10",
    )
    ws_accum.cell(row=r, column=7, value=f"=C{r} + D{r} + E{r} + F{r}")
    ws_accum.cell(
        row=r,
        column=8,
        value=f"=G{r} / ((1 + '02 — Assumptions'!B12) ^ B{r})",
    )

    for col_num in range(1, 9):
      c = ws_accum.cell(row=r, column=col_num)
      c.border = thin_border
      c.number_format = "$#,##0" if col_num not in [1, 2] else "General"
      if col_num in [1, 2]:
        c.alignment = Alignment(horizontal="center")

  # Tab 1: Dashboard
  ws_dash = wb["01 — Dashboard"]
  ws_dash.views.sheetView[0].showGridLines = True
  ws_dash["A1"] = "RETIREMENT & PERSONAL INVESTMENT DASHBOARD"
  ws_dash["A1"].font = Font(name="Calibri", size=16, bold=True, color="002060")

  kpis = [
      ("Current Portfolio", "='02 — Assumptions'!B5", "$#,##0"),
      ("Projected Portfolio", f"='03 — Accumulation'!G{years+1}", "$#,##0"),
      (
          "Required Nest Egg",
          "=(('02 — Assumptions'!B11 - '02 — Assumptions'!B12)) / 0.04",
          "$#,##0",
      ),
      ("Funding Ratio", "=B4/B5", "0.0%"),
  ]
  for idx, (label, formula, fmt) in enumerate(kpis, start=3):
    ws_dash.cell(row=idx, column=1, value=label).font = bold_font
    c = ws_dash.cell(row=idx, column=2, value=formula)
    c.font, c.number_format = bold_font, fmt

  for sheet in wb.worksheets:
    for col in sheet.columns:
      max_len = max(len(str(cell.value or "")) for cell in col)
      col_letter = get_column_letter(col[0].column)
      sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

  # OpenPyXL requires every cell number_format to be a string.
  # Normalize any accidental None values before saving.
  for sheet in wb.worksheets:
    for row in sheet.iter_rows():
      for cell in row:
        if cell.number_format is None:
          cell.number_format = "General"

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)
  return output


# ==========================================
# 1. SIDEBAR - USER INPUTS
# ==========================================
st.sidebar.markdown("""<div class="sidebar-brand">RETIREMENT<span>.</span></div><div class="sidebar-caption">Personal Investment Model</div>""", unsafe_allow_html=True)

st.sidebar.markdown('<div class="sidebar-section">Personal</div>', unsafe_allow_html=True)
current_age = st.sidebar.number_input(
    "Current Age", min_value=18, max_value=80, value=25
)
retirement_age = st.sidebar.number_input(
    "Retirement Age", min_value=current_age + 1, max_value=90, value=65
)
life_expectancy = st.sidebar.number_input(
    "Life Expectancy", min_value=retirement_age + 1, max_value=110, value=90
)

st.sidebar.markdown('<div class="sidebar-section">Portfolio & Savings</div>', unsafe_allow_html=True)
current_portfolio = st.sidebar.number_input(
    "Current Portfolio ($)", min_value=0.0, value=15000.0, step=1000.0
)
annual_income = st.sidebar.number_input(
    "Annual Income ($)", min_value=0.0, value=85000.0, step=2500.0
)
annual_contribution = st.sidebar.number_input(
    "Annual Contribution ($)", min_value=0.0, value=8000.0, step=500.0
)
employer_match_pct = (
    st.sidebar.slider("Employer Match (%)", 0.0, 15.0, 4.0) / 100.0
)
contribution_growth = (
    st.sidebar.slider("Annual Raise on Savings (%)", 0.0, 10.0, 2.0) / 100.0
)

st.sidebar.markdown('<div class="sidebar-section">Market Assumptions</div>', unsafe_allow_html=True)
expected_return = (
    st.sidebar.slider("Expected Annual Return (%)", 1.0, 15.0, 8.0) / 100.0
)
volatility = (
    st.sidebar.slider("Portfolio Volatility / Std Dev (%)", 1.0, 30.0, 15.0)
    / 100.0
)
inflation = (
    st.sidebar.slider("Expected Inflation Rate (%)", 0.0, 10.0, 3.0) / 100.0
)

st.sidebar.markdown('<div class="sidebar-section">Retirement Spending</div>', unsafe_allow_html=True)
retirement_spending = st.sidebar.number_input(
    "Desired Spending Goal ($/yr)", min_value=0.0, value=60000.0, step=2500.0
)
social_security = st.sidebar.number_input(
    "Social Security ($/yr)", min_value=0.0, value=20000.0, step=1000.0
)


# ==========================================
# EXCEL DOWNLOAD BUTTON IN SIDEBAR
# ==========================================

st.sidebar.divider()
#st.sidebar.markdown('<div class="sidebar-section">Export</div>', unsafe_allow_html=True)

current_inputs = {
    "current_age": current_age,
    "retirement_age": retirement_age,
    "life_expectancy": life_expectancy,
    "current_portfolio": current_portfolio,
    "annual_income": annual_income,
    "annual_contribution": annual_contribution,
    "employer_match_pct": employer_match_pct,
    "expected_return": expected_return,
    "volatility": volatility,
    "inflation": inflation,
    "retirement_spending": retirement_spending,
    "social_security": social_security,
    "contribution_growth": contribution_growth,
}

excel_data = build_excel_in_memory(current_inputs)
 
#st.sidebar.download_button(
    #label="Download Excel Model",
    #data=excel_data,
    #file_name="Retirement_Investment_Planner.xlsx",
    #mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#)


# ==========================================
# 2. CALCULATION ENGINE
# ==========================================
years_to_retire = retirement_age - current_age

years = np.arange(1, years_to_retire + 1)
ages = np.arange(current_age, retirement_age)

starting_bal = np.zeros(years_to_retire)
user_contribs = np.zeros(years_to_retire)
emp_matches = np.zeros(years_to_retire)
gains = np.zeros(years_to_retire)
ending_bal = np.zeros(years_to_retire)
real_vals = np.zeros(years_to_retire)

curr_bal = current_portfolio

for i in range(years_to_retire):
  starting_bal[i] = curr_bal
  user_contribs[i] = annual_contribution * ((1 + contribution_growth) ** i)
  emp_matches[i] = (
      annual_income * employer_match_pct * ((1 + contribution_growth) ** i)
  )

  tot_contrib = user_contribs[i] + emp_matches[i]
  gain = (starting_bal[i] + tot_contrib / 2) * expected_return
  gains[i] = gain

  curr_bal = starting_bal[i] + tot_contrib + gain
  ending_bal[i] = curr_bal
  real_vals[i] = curr_bal / ((1 + inflation) ** years[i])

df_accum = pd.DataFrame({
    "Age": ages,
    "Year": years,
    "Starting Balance": starting_bal,
    "User Contribution": user_contribs,
    "Employer Match": emp_matches,
    "Investment Gain": gains,
    "Ending Portfolio": ending_bal,
    "Real Purchasing Power": real_vals,
})

# Readiness Calculations
projected_assets = (
    ending_bal[-1] if len(ending_bal) > 0 else current_portfolio
)
net_spending = max(0.0, retirement_spending - social_security)
required_assets = net_spending / 0.04  # 4% Rule
funding_ratio = (
    (projected_assets / required_assets) if required_assets > 0 else 1.0
)

if funding_ratio < 0.80:
  status_color = "red"
  status_label = "At Risk"
elif funding_ratio < 1.00:
  status_color = "orange"
  status_label = "Needs Improvement"
elif funding_ratio < 1.20:
  status_color = "blue"
  status_label = "On Track"
else:
  status_color = "green"
  status_label = "Strong"


# ==========================================
# 3. DASHBOARD KPI METRICS
# ==========================================
st.markdown("""<div class="section-header"><div class="section-title">Model Snapshot</div><div class="section-meta">Current assumptions</div></div>""", unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)

status_class = {"red":"status-risk", "orange":"status-warning", "blue":"status-track", "green":"status-strong"}.get(status_color, "status-track")

def kpi_card(label, value, note, css_class=""):
    return f"""<div class=\"kpi-card\"><div class=\"kpi-label\">{label}</div><div class=\"kpi-value {css_class}\">{value}</div><div class=\"kpi-note\">{note}</div></div>"""

with col1:
    st.markdown(kpi_card("Projected Nest Egg", f"${projected_assets:,.0f}", f"At age {retirement_age}"), unsafe_allow_html=True)
with col2:
    st.markdown(kpi_card("Required Nest Egg", f"${required_assets:,.0f}", "4% withdrawal framework"), unsafe_allow_html=True)
with col3:
    st.markdown(kpi_card("Funding Ratio", f"{funding_ratio*100:.1f}%", "Projected / Required"), unsafe_allow_html=True)
with col4:
    st.markdown(kpi_card("Planning Status", status_label.upper(), "Based on current assumptions", status_class), unsafe_allow_html=True)

st.markdown("""<div class="section-header"><div class="section-title">Portfolio Growth</div><div class="section-meta">Accumulation period</div></div>""", unsafe_allow_html=True)


# ==========================================
# 4. TABBED NAVIGATION
# ==========================================
tab1, tab2, tab3, tab4 = st.tabs([
    "Growth Projection",
    "Scenario Analysis",
    "Monte Carlo",
    "Full Schedule",
])

with tab1:
  st.markdown("""<div class="section-header"><div class="section-title">Nominal vs. Real Portfolio Value</div><div class="section-meta">Purchasing power</div></div>""", unsafe_allow_html=True)

  fig, ax = plt.subplots(figsize=(10, 4))
  ax.plot(
      df_accum["Age"],
      df_accum["Ending Portfolio"] / 1e6,
      label="Nominal Portfolio ($M)",
      color="#2f8053",
      linewidth=2,
  )
  ax.plot(
      df_accum["Age"],
      df_accum["Real Purchasing Power"] / 1e6,
      label="Inflation-Adjusted ($M)",
      color="#7d8982",
      linestyle="--",
  )
  ax.set_xlabel("Age")
  ax.set_ylabel("Portfolio Value ($ Millions)")
  ax.grid(True, alpha=0.3)
  ax.legend()
  st.pyplot(fig)

  tot_user_contrib = df_accum["User Contribution"].sum()
  tot_emp_match = df_accum["Employer Match"].sum()
  tot_gains = df_accum["Investment Gain"].sum()

  st.markdown("""<div class="section-header"><div class="section-title">Portfolio Composition</div><div class="section-meta">Sources of projected growth</div></div>""", unsafe_allow_html=True)
  col_a, col_b = st.columns([1, 2])
  with col_a:
    st.write(f"**Initial Portfolio:** ${current_portfolio:,.0f}")
    st.write(f"**Total User Savings:** ${tot_user_contrib:,.0f}")
    st.write(f"**Total Employer Match:** ${tot_emp_match:,.0f}")
    st.write(f"**Total Investment Gains:** ${tot_gains:,.0f}")
  with col_b:
    fig_pie, ax_pie = plt.subplots(figsize=(5, 3))
    ax_pie.pie(
        [current_portfolio + tot_user_contrib, tot_emp_match, tot_gains],
        labels=["Your Contributions", "Employer Match", "Compound Growth"],
        autopct="%1.1f%%",
        startangle=140,
        colors=["#2f8053", "#7d8982", "#b8c4bd"],
    )
    st.pyplot(fig_pie)

with tab2:
  st.markdown("""<div class="section-header"><div class="section-title">Market Return Sensitivity</div><div class="section-meta">Scenario analysis</div></div>""", unsafe_allow_html=True)
  rates = [0.05, 0.08, 0.10]
  scen_names = ["Conservative (5%)", "Base Case (8%)", "Aggressive (10%)"]
  scen_results = []

  for r in rates:
    c_bal = current_portfolio
    for i in range(years_to_retire):
      tc = (annual_contribution + annual_income * employer_match_pct) * (
          (1 + contribution_growth) ** i
      )
      c_bal = (c_bal + tc / 2) * (1 + r) + tc / 2
    scen_results.append(c_bal)

  df_scen = pd.DataFrame(
      {"Scenario": scen_names, "Projected Portfolio": scen_results}
  )
  df_scen["Projected Portfolio"] = df_scen["Projected Portfolio"].map(
      "${:,.0f}".format
  )
  st.table(df_scen)

with tab3:
  st.markdown("""<div class="section-header"><div class="section-title">Monte Carlo Risk Analysis</div><div class="section-meta">Stochastic simulation</div></div>""", unsafe_allow_html=True)
  st.markdown(
      "Simulates sequence-of-returns risk using a Gaussian distribution based"
      " on expected returns and volatility."
  )

  if st.button("🚀 Run Monte Carlo Simulation"):
    num_sims = 5000
    sim_returns = np.random.normal(
        expected_return, volatility, (num_sims, years_to_retire)
    )
    paths = np.zeros((num_sims, years_to_retire + 1))
    paths[:, 0] = current_portfolio

    for yr in range(1, years_to_retire + 1):
      tc = (annual_contribution + annual_income * employer_match_pct) * (
          (1 + contribution_growth) ** (yr - 1)
      )
      paths[:, yr] = (paths[:, yr - 1] + tc) * (1 + sim_returns[:, yr - 1])

    final_vals = paths[:, -1]

    p10 = np.percentile(final_vals, 10)
    p25 = np.percentile(final_vals, 25)
    p50 = np.percentile(final_vals, 50)
    p75 = np.percentile(final_vals, 75)
    p90 = np.percentile(final_vals, 90)

    prob_success = np.mean(final_vals >= required_assets) * 100

    m_col1, m_col2 = st.columns([1, 2])
    with m_col1:
      st.metric("Probability of Meeting Goal", f"{prob_success:.1f}%")
      st.write(f"**10th Percentile:** ${p10:,.0f}")
      st.write(f"**25th Percentile:** ${p25:,.0f}")
      st.write(f"**50th (Median):** ${p50:,.0f}")
      st.write(f"**75th Percentile:** ${p75:,.0f}")
      st.write(f"**90th Percentile:** ${p90:,.0f}")

    with m_col2:
      fig_mc, ax_mc = plt.subplots(figsize=(8, 4))
      for i in range(min(100, num_sims)):
        ax_mc.plot(
            np.arange(current_age, retirement_age + 1),
            paths[i, :] / 1e6,
            color="#6f7c74",
            alpha=0.1,
        )
      ax_mc.set_title("Simulated Portfolio Trajectories (First 100 Paths)")
      ax_mc.set_xlabel("Age")
      ax_mc.set_ylabel("Portfolio Value ($M)")
      ax_mc.grid(True, alpha=0.3)
      st.pyplot(fig_mc)

with tab4:
  st.markdown("""<div class="section-header"><div class="section-title">Year-by-Year Schedule</div><div class="section-meta">Accumulation detail</div></div>""", unsafe_allow_html=True)
  formatted_df = df_accum.copy()
  for col in [
      "Starting Balance",
      "User Contribution",
      "Employer Match",
      "Investment Gain",
      "Ending Portfolio",
      "Real Purchasing Power",
  ]:
    formatted_df[col] = formatted_df[col].map("${:,.2f}".format)

  st.dataframe(formatted_df, use_container_width=True)

# ============================================================
# FOOTER
# ============================================================

st.markdown(
    """<div class="app-footer"><span>JOEL · FINANCE & ANALYTICS</span><span>RETIREMENT PLANNING MODEL · FOR EDUCATIONAL USE</span></div>""",
    unsafe_allow_html=True,
)
